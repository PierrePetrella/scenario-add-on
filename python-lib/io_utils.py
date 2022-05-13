# -*- coding: utf-8 -*-
import dataiku
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from tempfile import NamedTemporaryFile


DUMMY_CONSTANT = "foo"

# Init project_var to value
def set_project_var (project_var, value):
    client = dataiku.api_client()
    project_api = client.get_default_project()
    v = project_api.get_variables()
    v["standard"][project_var] = value
    project_api.set_variables(v)

# Increment project_var by inc
def inc_project_var(project_var, inc = 1):
    client = dataiku.api_client()
    project_api = client.get_default_project()
    v = project_api.get_variables()
    v["standard"][project_var] = v["standard"][project_var] + inc
    project_api.set_variables(v)

def write_wb_to_managed_folder(managed_folder_handle, wb, file_path):
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        output = tmp.read()
        with managed_folder_handle.get_writer(file_path + ".xlsx") as w:
            w.write(output)

def write_df_in_wb(df):
    rows = dataframe_to_rows(df, index=False)
    wb = openpyxl.Workbook()
    ws = wb.active
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
             ws.cell(row=r_idx, column=c_idx, value=value)
    return wb

def write_csv_to_managed_folder (managed_folder_handle, df, file_path):
    with managed_folder_handle.get_writer(file_path +".csv") as writer:
        writer.write(df.to_csv().encode("utf-8"))

def read_csv_from_managed_folder (handle_managed_folder, file_path):
    with handle_managed_folder.get_download_stream(file_path) as f:
        df = pd.read_csv(f)
    return df
        
# Read excel from Managed Folder
def read_excel_from_managed_folder(managed_folder_handle,file_path):
    with managed_folder_handle.get_download_stream(file_path) as f:
        bytes_in = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(bytes_in)
    return wb

#Write wb to a Managed Folder
def write_wb_to_managed_folder(managed_folder_handle,wb, file_path):
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        output = tmp.read()
        with output_folder.get_writer(file_path) as w:
            w.write(output)


def rename_managed_folder_file(handle_managed_folder, file_path, new_file_path):
    with handle_managed_folder.get_download_stream(file_path) as f:
        file = f.read()
    with handle_managed_folder.get_writer(new_file_path) as w:
        w.write(file)
    handle_managed_folder.delete_path(file_path)